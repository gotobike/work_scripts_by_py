#!/usr/bin/env python
# @author: Kyle Qin
# @email: kyleqinmail@gmail.com
# @version: v2.0
# @description: h3c光模块巡检信息自动化导出

"""
1.脚本与源文件同目录，并且目录下无其他文件夹
2.第三方依赖：pip install parse ,pip install openpyxl
3.当前目录下cmd运行：python optical_modules_inspection.py
"""
import time
import datetime
import glob
import os
from parse import compile
import openpyxl as opxl


# pysnooper为测试模块，当程序内部出现问题时可取消注释，然后取消异常函数前的装饰器注释即可
# import pysnooper


# @pysnooper.snoop()
def xlsx_file_create(header: tuple) -> str:
    """
    创建xlsx持久化文件
    :param header: xlsx header title
    :return: xlsx_path
    """
    wb = opxl.Workbook()
    ws = wb.active
    ws.title = "Optical modules"

    for i in range(len(header)):
        ws.cell(row=1, column=i + 1).value = header[i]

    xlsx_path = f"{os.getcwd()}\光模块信息导出_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


# @pysnooper.snoop()
def get_resource() -> tuple:
    """
    源文件路径查找
    :return: resource_path,equ_ipa_dict
    """
    resource_path = []
    equ_ipa_dict = {}
    pwd = os.getcwd()
    pattern_equ_ipa = compile("{}({}).log")
    for root, dirs_name, files_name in os.walk(pwd):
        if len(dirs_name) == 0:
            for file_name in files_name:
                if file_name.endswith(".log"):
                    equ_ipa_dict.update(
                        {pattern_equ_ipa.parse(file_name)[0]: pattern_equ_ipa.parse(file_name)[1]})
                    resource_path.append(f"{root}\{file_name}")
    return tuple(resource_path), equ_ipa_dict


# @pysnooper.snoop()
def get_resource_glob() -> tuple:
    """
    源文件路径查找
    :return: resource_path,equ_ipa_dict
    """
    resource_path = []
    equ_ipa_dict = {}
    pwd = os.getcwd()
    pattern_equ_ipa = compile("{}({}).log")
    for file_name in glob.glob("*.log"):
        equ_ipa_dict.update({pattern_equ_ipa.parse(file_name)[0]: pattern_equ_ipa.parse(file_name)[1]})
        resource_path.append(f"{pwd}\{file_name}")
    print(resource_path,equ_ipa_dict)
    return tuple(resource_path), equ_ipa_dict


# @pysnooper.snoop()
def optical_module_export(file_path: str) -> dict:
    """
    光模块信息解析函数
    :param file_path: 单个文件路径
    :return: {设备名:{端口名:[(rx,tx),],}}
    """
    equ_name = ""
    flag_port = False
    flag_channel = False
    port_val_dict = {}
    val_channels = []
    equ_cmd_pattern = compile("<{}>")
    port_equ_pattern = compile("{} transceiver diagnostic information:")

    # replaced：fileinput.input()
    with open(file_path, encoding="utf-8") as f:
        for line in f.readlines():
            line = line.strip()
            # 此处应判断作用区域...
            if line.startswith("<"):
                equ_cmd = equ_cmd_pattern.parse(line)
                if equ_cmd != None:
                    equ_name = equ_cmd_pattern.parse(line)[0]

            if line.endswith("transceiver diagnostic information:"):
                port_name = port_equ_pattern.parse(line)[0]
                flag_port = True
            # port judge:
            if flag_port:
                if line.startswith("Alarm thresholds:"):
                    flag_channel = False
                    flag_port = False
                    port_val_dict.update({port_name: val_channels})
                    val_channels = []

                if flag_channel:
                    channel_list = line.split()
                    val_channels.append((channel_list[-2], channel_list[-1]))

                if line.endswith("TX power(dBm)"):
                    flag_channel = True

    return {equ_name: port_val_dict}


# @pysnooper.snoop()
def save_to_excel(xlsx_path: str, files_parse_dict: dict) -> None:
    """
    解析出的信息持久化到xlsx文件中
    :param xlsx_path: 文件路径
    :param files_parse_dict: 解析文件
    :return:
    """
    files_parse_dict_keys = files_parse_dict.keys()
    try:
        wb = opxl.load_workbook(xlsx_path)
        ws = wb.worksheets[0]
        for key_equ in files_parse_dict_keys:
            ipad = equ_ipa_dict.get(key_equ)
            port_dict = files_parse_dict.get(key_equ)
            port_dict_keys = port_dict.keys()
            current_rows = ws.max_row
            ws.cell(row=current_rows + 1, column=1).value = ipad
            ws.cell(row=current_rows + 1, column=2).value = key_equ
            for j in range(len(port_dict_keys)):
                channel_list = port_dict.get(list(port_dict_keys)[j])
                if j == 0:
                    start_pos = ws.max_row
                else:
                    start_pos = ws.max_row + 1
                count = 0
                for i in range(start_pos, start_pos + len(channel_list)):
                    if count == 0:
                        ws.cell(row=i, column=3).value = list(port_dict_keys)[j]
                    ws.cell(row=i, column=4).value = count + 1
                    ws.cell(row=i, column=5).value = channel_list[count][0]
                    ws.cell(row=i, column=6).value = channel_list[count][1]
                    count = count + 1
        wb.save(xlsx_path)
    except Exception as e:
        print(e)
    finally:
        wb.close()
    return None


# @pysnooper.snoop()
def deal_files_parse(files_parse_dict: dict) -> dict:
    """
    解析文件筛选函数
    :param files_parse_dict:
    :return:
    """
    equ_keys = files_parse_dict.keys()
    for equ_name in equ_keys:
        port_keys = files_parse_dict.get(equ_name)
        port_name_dict = files_parse_dict[equ_name]
        # find bit_map
        bit_map = []
        for port_name in port_keys:
            if port_name.startswith("Hund"):
                bit_map.append(100)
            if port_name.startswith("Ten"):
                bit_map.append(10)
            if port_name.startswith("Tw"):
                bit_map.append(25)
        # judge bit_map
        bit_map_set = set(bit_map)
        val =""

        match len(bit_map_set):
            case 1:
                continue
            case 2:
                val_1, val_2 = bit_map_set
                val = max(val_2, val_1)
            case 3:
                val1, val2, val3 = bit_map_set
                val = max(val1, val2, val3)

        for i in range(len(bit_map)):
            if bit_map[i] == val:
                continue
            del port_name_dict[list(port_keys)[i]]

    return files_parse_dict


if __name__ == "__main__":
    start_time = time.time()
    header = ("IP ADDRES", "设备名", "端口号", "channel", "H3C Rx", "H3C Tx")
    xlsx_path = xlsx_file_create(header)
    files_parse_dict = {}
    files_path, equ_ipa_dict = get_resource()
    files_path, equ_ipa_dict = get_resource_glob()
    files_num = len(files_path)
    print(f"此次执行，源文件共计{files_num}个...解析中，请稍后-/*")
    print("#" * 10)
    for i in range(files_num):
        files_parse_dict.update(optical_module_export(files_path[i]))
    files_parse_dict = deal_files_parse(files_parse_dict)
    save_to_excel(xlsx_path, files_parse_dict)
    print(f"本次导出xlsx文件路径：{xlsx_path}")
    end_time = time.time()
    print(f"运行结束！共耗时{(end_time - start_time):.2f}s")
