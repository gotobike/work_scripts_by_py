#!/usr/bin/env python
# @author: Kyle Qin
# @email: kyleqinmail@gmail.com
# @version: v1.1
# @description: juniper防火墙策略整理

import glob
import fileinput
import datetime
import time
import copy
import parse
import openpyxl as opxl
# 测试模块，非必须
import pysnooper

# pattern_global
pattern_policy_orient_id_desc = parse.compile(
    "set security policies from-zone {} to-zone {} policy {} description {}")
pattern_policy_source_address = parse.compile(
    "set security policies from-zone {} to-zone {} policy {} match source-address {}")
pattern_policy_destination_address = parse.compile(
    "set security policies from-zone {} to-zone {} policy {} match destination-address {}")
pattern_policy_application = parse.compile(
    "set security policies from-zone {} to-zone {} policy {} match application {}")
pattern_policy_then = parse.compile(
    "set security policies from-zone {} to-zone {} policy {} then {}")

pattern_zones_application = parse.compile(
    "set security zones security-zone {} address-book address-set {} {} {}")

pattern_applications_application_set_1 = parse.compile("set applications application-set {} application {}")
pattern_applications_application_set_2 = parse.compile("set applications application-set {} application-set {}")


# pattern_applications_application = parse.compile(
#     "set applications application {} destination-port {}")


def get_resource() -> tuple:
    """
    获取文件路径
    :return:
    """
    files_path = glob.glob("*_cfg.txt")
    return (files_path,)


# def add_set_like_optimize(add_set: dict) -> dict:
#     for add_set_key in add_set.keys():
#         for add_set_ip in add_set[add_set_key]:
#             if add_set.get(add_set_ip, "-1") != "-1":
#                 add_set[add_set_key].extend(add_set[add_set_ip])
#                 add_set[add_set_key].remove(add_set_ip)
#     return add_set
#
#
def resource_list_optimize(resource_list: list, address_set: dict, application_set: dict) -> tuple:
    # while True:需要查找算法
    # 需要优化
    # for i in range(8):
    #     address_set = add_set_like_optimize(address_set)
    # for i in range(8):
    #     application_set = add_set_like_optimize(application_set)

    for resource_dict in resource_list:
        # Inside
        dict_key_tuple = ("Inside", "Outside", "ports")
        for dict_key in dict_key_tuple:
            for resource_dict_inside in resource_dict[dict_key]:
                if address_set.get(resource_dict_inside, "-1") != "-1":
                    resource_dict[dict_key].extend(address_set[resource_dict_inside])
                    resource_dict[dict_key].remove(resource_dict_inside)

    return tuple(resource_list)


# @pysnooper.snoop()
def juniper_firewall_policy(files_path: tuple, dict_mod: dict) -> tuple:
    resource_list = []
    address_set = {}
    address_ip = {}
    application_set = {}
    area = ""
    dict = copy.deepcopy(dict_mod)
    files_path = list(*files_path)

    for line in fileinput.input(files_path[0], encoding="utf-8"):
        line = line.strip()

        if line.startswith("C"):
            area = "SH"
        if line.startswith("H"):
            area = "HF"

        dict["area"] = area

        if line.startswith("set security policies"):
            if line.rfind("description") > 0:
                orient_left, orient_right, policy_id, policy_desc = pattern_policy_orient_id_desc.parse(line)
                dict["orient"][0] = orient_left
                dict["orient"][1] = orient_right
                dict["policy_id"] = policy_id
                dict["desc"] = policy_desc
            if line.rfind("source-address") > 0:
                _, _, _, source_address = pattern_policy_source_address.parse(line)
                dict["Inside"].append(source_address)
            if line.rfind("destination-address") > 0:
                _, _, _, dest_address = pattern_policy_destination_address.parse(line)
                dict["Outside"].append(dest_address)
            if line.rfind("application") > 0:
                _, _, _, app_ports = pattern_policy_application.parse(line)
                dict["ports"].append(app_ports)
            if line.rfind("then") > 0:
                _, _, _, services = pattern_policy_then.parse(line)
                if dict["policy_id"] != "":
                    resource_list.append(dict)
                    dict = copy.deepcopy(dict_mod)

        if line.startswith("set security zones security-zone"):
            if line.rfind("address-set") > 0:
                _, ip_group, add_flag, add_ip_group = pattern_zones_application.parse(line)
                if add_flag == "address-set":
                    if address_set.get(ip_group, "-1") != "-1":
                        address_set[ip_group].append(add_ip_group)
                    else:
                        address_set.update({ip_group: [add_ip_group]})
                else:
                    if address_ip.get(ip_group, "-1") != "-1":
                        address_ip[ip_group].append(add_ip_group)
                    else:
                        address_ip.update({ip_group: [add_ip_group]})

        if line.startswith("set applications application-set"):
            if line.rfind("application") > line.rfind("application-set"):
                app_set, app_group = pattern_applications_application_set_1.parse(line)
                if application_set.get(app_set, "-1") != "-1":
                    application_set[app_set].append(app_group)
                else:
                    application_set.update({app_set: [app_group]})
            else:
                app_set, app_group = pattern_applications_application_set_2.parse(line)
                if application_set.get(app_set, "-1") != "-1":
                    application_set[app_set].append(app_group)
                else:
                    application_set.update({app_set: [app_group]})

    address_set = address_set_optimize(address_set, address_ip)

    # print(address_set)
    # print(address_ip)

    return (resource_list, address_set, application_set)


# @pysnooper.snoop()
def address_set_optimize(address_set: dict, address_ip: dict) -> dict:
    address_set = address_set
    # print(address_ip)
    # print("*"*15)
    # 判断ip——group和ip地址的对应关系

    for address_set_key in address_set.keys():
        for add_set_group in address_set[address_set_key]:
            if address_set.get(add_set_group, "-1") != "-1":
                address_set[address_set_key].extend(address_set[add_set_group])
                address_set[address_set_key].remove(add_set_group)
    for i in range(4):
        for address_set_key in address_set.keys():
            for add_set_group in address_set[address_set_key]:
                if address_ip.get(add_set_group,"-1") != "-1":
                    address_set[address_set_key].extend(address_ip[add_set_group])
                    address_set[address_set_key].remove(add_set_group)

    return address_set


# @pysnooper.snoop()
def xlsx_file_create(header: tuple) -> str:
    """
    创建xlsx持久化文件
    :param header: xlsx header title
    :return: xlsx_path
    """
    wb = opxl.Workbook()
    ws0 = wb.active
    ws0.title = "合肥内到外"
    ws1 = wb.create_sheet()
    ws1.title = "合肥外到内"
    ws2 = wb.create_sheet()
    ws2.title = "上海内到外"
    ws3 = wb.create_sheet()
    ws3.title = "上海外到内"

    for worksheet in wb.worksheets:
        for i in range(len(header)):
            worksheet.cell(row=1, column=i + 1).value = header[i]

    xlsx_path = f".\数据中心防护防火墙策略ID-系统对应表_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(xlsx_path)
    return xlsx_path


def save_to_excel(xlsx_path: str):
    try:
        wb = opxl.load_workbook(xlsx_path)
        ws_list = wb.worksheets
        print(ws_list)
    except Exception as e:
        print(e)
    finally:
        wb.close()


if __name__ == "__main__":
    start_time = time.time()
    header = ("序号", "策略id", "源ip", "目的ip", "端口", "描述")
    # area:SH or HF ;orient:[outside,inside] ; etc...
    policy_dict_mod = {"area": "", "orient": ["", ""], "policy_id": "",
                       "Inside": [], "Outside": [], "ports": [], "desc": ""}
    # xlsx_path = xlsx_file_create(header)
    resource_files = get_resource()
    files_num = len(resource_files[0])
    print(f"此次执行，源文件共计{files_num}个...解析中，请稍后-/*")
    print("#" * 10)
    resource_list, address_set, application_set = juniper_firewall_policy(resource_files, policy_dict_mod)

    resource_tuple = resource_list_optimize(resource_list, address_set, application_set)
    print(resource_tuple)

    # print(f"本次导出xlsx文件路径：{xlsx_path}")
    end_time = time.time()
    print(f"运行结束！共耗时{(end_time - start_time):.2f}s")
